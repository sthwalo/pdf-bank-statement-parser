#!/usr/bin/env python3
"""
Audit script to identify uncategorized transactions in the cashbook.
This helps identify transactions that need to be properly categorized.
"""

import os
import pandas as pd
import argparse
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class TransactionCategorizer:
    def __init__(self):
        # Define specific transaction patterns and their categories
        self.specific_patterns = {
            # Bank Charges
            'bank charges': 'Bank Charges',
            'fees': 'Bank Charges',
            'account fee': 'Bank Charges',
            'item fee': 'Bank Charges',
            'manual reversal fee': 'Bank Charges',
            'unsuccessful f declined': 'Bank Charges',
            'unpaid no funds': 'Bank Charges',
            'fee tcib': 'Bank Charges',
            'swift commission': 'Bank Charges',
            'replacement fee': 'Bank Charges',
            
            # Vehicle Related
            'car hire': 'Vehicle Hire',
            'car rental': 'Vehicle Hire',
            'truck hire': 'Vehicle Hire',
            'quantum hire': 'Vehicle Hire',
            'rentals': 'Vehicle Hire',
            'outward swift r024': 'Vehicle Hire',
            
            # Utilities
            'electricity prepaid': 'Electricity',
            
            # Education
            'computer lessons': 'Education',
            'extra lessons': 'Education',
            'simphiwe': 'School Fees',
            
            # Household/Groceries
            'checkers': 'Household',
            'woolworths': 'Household',
            'pnp': 'Household',
            'spar': 'Household',
            'grocc': 'Household',
            'gro ': 'Household',
            'makro': 'Household',
            'food': 'Household',
            'markham': 'Household', 
            'edgars': 'Household',
            'furniture': 'Assets',

            
            # Toll Fees
            'plaza': 'Toll Fees',
            
            # Professional Services
            'luc trs': 'Outsourced Services',
            'mas': 'Outsourced Services',
            
            # Salaries
            'tendai': 'Salaries and Wages',
            'g gr gu': 'Salaries and Wages',
            
            # Stationery
            'stationery': 'Stationery',
            
            # Fuel
            'astron energy': 'Fuel',
            'petrol': 'Fuel',
            
            # Bond Payments
            'rental': 'Bond Payments'
        }
        
        # Define confidence levels
        self.confidence_levels = {
            'High': ['Bank Charges', 'Vehicle Hire', 'Electricity', 'Education', 
                    'School Fees', 'Toll Fees', 'Outsourced Services', 
                    'Salaries and Wages', 'Stationery', 'Fuel'],
            'Medium': ['Household'],
            'Low': ['General']
        }

    def suggest_category(self, description, amount):
        """Analyze transaction and suggest account type and category."""
        desc = str(description).lower()
        
        # Check specific patterns first
        for pattern, category in self.specific_patterns.items():
            if pattern in desc:
                confidence = next((level for level, cats in self.confidence_levels.items() 
                                if category in cats), 'Low')
                return {
                    'Account_Type': 'Expense',
                    'Category': category,
                    'Suggested_Account': f"{category} Expense",
                    'Confidence': confidence
                }
        
        # Default to General Expense if no specific match
        return {
            'Account_Type': 'Expense',
            'Category': 'General',
            'Suggested_Account': 'General Expense',
            'Confidence': 'Low'
        }

def suggest_categories(df):
    """Suggest categories based on transaction descriptions."""
    categorizer = TransactionCategorizer()
    
    suggestions = []
    for _, row in df.iterrows():
        suggestion = categorizer.suggest_category(row['description'], row['amount'])
        if suggestion:
            suggestions.append({
                'Date': row['date'],
                'Description': row['description'],
                'Amount': row['amount'],
                'Account_Type': suggestion['Account_Type'],
                'Category': suggestion['Category'],
                'Suggested_Account': suggestion['Suggested_Account'],
                'Confidence': suggestion['Confidence']
            })
    
    # Return empty DataFrame instead of None if no suggestions
    return pd.DataFrame(suggestions) if suggestions else pd.DataFrame(columns=[
        'Date', 'Description', 'Amount', 'Account_Type', 'Category', 
        'Suggested_Account', 'Confidence'
    ])

def audit_uncategorized(excel_path=None, csv_path=None, output_path=None):
    """Audit uncategorized transactions from Excel cashbook or CSV file."""
    
    if excel_path:
        print(f"Reading from Excel file: {excel_path}")
        df = pd.read_excel(excel_path, sheet_name='Detailed Transactions')
    elif csv_path:
        print(f"Reading from CSV file: {csv_path}")
        df = pd.read_csv(csv_path)
    else:
        raise ValueError("No input file specified")

    # Get uncategorized transactions
    uncategorized = df[df['Account'].str.contains('Uncategorized|Unknown', na=False)].copy()
    
    if len(uncategorized) == 0:
        print("No uncategorized transactions found!")
        return None
        
    # Sort by date and amount
    uncategorized = uncategorized.sort_values(['date', 'amount'])
    
    if output_path:
        # Save to Excel with formatting
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            uncategorized.to_excel(writer, index=False, sheet_name='Uncategorized')
            
            # Get the worksheet
            ws = writer.sheets['Uncategorized']
            
            # Format headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E6E6E6', 
                                      end_color='E6E6E6', 
                                      fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Format date column
            for cell in ws['A']:
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy-mm-dd'
                    
            # Format amount column with ZAR currency
            for cell in ws['C']:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = 'R #,##0.00;[Red]-R #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            
            # Adjust column widths
            for idx, col in enumerate(ws.columns, 1):
                ws.column_dimensions[get_column_letter(idx)].width = 20
                
        print(f"Audit report saved to: {output_path}")
        
    return uncategorized

def analyze_uncategorized_transactions(file_path):
    """Analyze uncategorized transactions and provide suggestions."""
    # Read the uncategorized transactions
    df = pd.read_excel(file_path)
    
    # Initialize categorizer
    categorizer = TransactionCategorizer()
    
    # Analyze each transaction
    suggestions = []
    for _, row in df.iterrows():
        suggestion = categorizer.suggest_category(row['description'], row['amount'])
        suggestions.append({
            'Date': row['date'],
            'Description': row['description'],
            'Amount': row['amount'],
            'Account_Type': suggestion['Account_Type'],
            'Category': suggestion['Category'],
            'Suggested_Account': suggestion['Suggested_Account'],
            'Confidence': suggestion['Confidence']
        })
    
    # Create suggestions DataFrame
    suggestions_df = pd.DataFrame(suggestions)
    
    # Save enhanced analysis
    output_path = 'uncategorized_analysis.xlsx'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        suggestions_df.to_excel(writer, index=False)
        
        # Format worksheet
        ws = writer.sheets['Sheet1']
        
        # Format headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Format columns
        for idx, col in enumerate(ws.columns, 1):
            # Set column width
            ws.column_dimensions[get_column_letter(idx)].width = 20
            
            # Format date and amount columns
            if idx == 1:  # Date column
                for cell in list(col)[1:]:  # Skip header
                    cell.number_format = 'yyyy-mm-dd'
            elif idx == 3:  # Amount column
                for cell in list(col)[1:]:  # Skip header
                    cell.number_format = 'R #,##0.00;[Red]-R #,##0.00'
                    cell.alignment = Alignment(horizontal='right')
    
    return suggestions_df

def main():
    """Main function to run the audit script."""
    parser = argparse.ArgumentParser(description='Audit uncategorized transactions in the cashbook.')
    parser.add_argument('--excel', help='Path to the Excel cashbook file')
    parser.add_argument('--csv', help='Path to the original CSV file')
    parser.add_argument('--output', help='Path to save the audit report')
    args = parser.parse_args()
    
    # If no arguments provided, prompt for input
    excel_path = args.excel
    csv_path = args.csv
    output_path = args.output
    
    if not excel_path and not csv_path:
        print("No input file specified. Please provide path to file:")
        file_path = input("Enter path to Excel cashbook or CSV file: ").strip()
        
        if file_path.endswith('.xlsx'):
            excel_path = file_path
        elif file_path.endswith('.csv'):
            csv_path = file_path
        else:
            print(f"Unsupported file format: {file_path}")
            return
    
    if not output_path:
        default_output = "uncategorized_audit.xlsx"
        use_default = input(f"Save audit report to {default_output}? (y/n): ").strip().lower()
        if use_default == 'y':
            output_path = default_output
    
    # Run the audit
    uncategorized = audit_uncategorized(excel_path, csv_path, output_path)
    
    # Suggest categories
    if uncategorized is not None and len(uncategorized) > 0:
        suggestions = suggest_categories(uncategorized)
        
        # Save suggestions to the output file if it exists
        if output_path and os.path.exists(output_path):
            with pd.ExcelWriter(output_path, mode='a') as writer:
                suggestions.to_excel(writer, sheet_name='Suggestions', index=False)
            print(f"Category suggestions added to: {output_path}")
    
    input_file = output_path
    print(f"Analyzing transactions from {input_file}...")
    results = analyze_uncategorized_transactions(input_file)
    print("\nAnalysis complete! Results saved to 'uncategorized_analysis.xlsx'")
    print(f"\nFound {len(results)} transactions to categorize")
    print("\nConfidence levels summary:")
    print(results['Confidence'].value_counts())

if __name__ == "__main__":
    main()
