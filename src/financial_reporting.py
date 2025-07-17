import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

def read_cashbook_data(cashbook_path):
    """
    Read data from the Annual Cashbook Excel file.
    
    Args:
        cashbook_path (str): Path to the Annual Cashbook Excel file
        
    Returns:
        dict: Dictionary containing DataFrames for each sheet in the cashbook
    """
    print(f"Reading cashbook data from: {cashbook_path}")
    
    # Read all sheets from the cashbook
    cashbook_data = {}
    
    try:
        # Read the Trial Balance sheet
        trial_balance = pd.read_excel(cashbook_path, sheet_name='Trial Balance')
        cashbook_data['trial_balance'] = trial_balance
        
        # Read the Monthly Summary sheet
        monthly_summary = pd.read_excel(cashbook_path, sheet_name='Monthly Summary')
        cashbook_data['summary'] = monthly_summary
        
        # Read the Detailed Transactions sheet
        details = pd.read_excel(cashbook_path, sheet_name='Detailed Transactions')
        cashbook_data['details'] = details
        
        print(f"Cashbook data loaded successfully. Found sheets: {list(cashbook_data.keys())}")
        return cashbook_data
        
    except Exception as e:
        print(f"Error reading cashbook data: {str(e)}")
        print("Available sheets in the workbook:")
        xls = pd.ExcelFile(cashbook_path)
        print(xls.sheet_names)
        raise

def generate_income_statement(cashbook_data):
    """Generate Income Statement from trial balance data."""
    print("Generating Income Statement...")
    
    # Get trial balance data
    trial_balance = cashbook_data['trial_balance']
    
    # Filter for Income and Expense accounts
    income_accounts = trial_balance[trial_balance['Account Type'].str.contains('Income', na=False)].copy()
    expense_accounts = trial_balance[trial_balance['Account Type'].str.contains('Expense', na=False)].copy()
    
    # Create income statement DataFrame
    income_statement = []
    
    # Add income section
    income_statement.append({
        'Account': 'INCOME',
        'Amount': None
    })
    
    # Add income accounts
    total_income = 0
    for _, row in income_accounts.iterrows():
        net_amount = row['Credit'] - row['Debit']
        total_income += net_amount
        income_statement.append({
            'Account': row['Account'],
            'Amount': net_amount
        })
    
    # Add total income
    income_statement.append({
        'Account': 'Total Income',
        'Amount': total_income
    })
    
    # Add expenses section
    income_statement.append({
        'Account': 'EXPENSES',
        'Amount': None
    })
    
    # Add expense accounts
    total_expenses = 0
    for _, row in expense_accounts.iterrows():
        net_amount = row['Debit'] - row['Credit']
        total_expenses += net_amount
        income_statement.append({
            'Account': row['Account'],
            'Amount': net_amount
        })
    
    # Add total expenses
    income_statement.append({
        'Account': 'Total Expenses',
        'Amount': total_expenses
    })
    
    # Calculate and add net profit/loss
    net_profit = total_income - total_expenses
    income_statement.append({
        'Account': 'Net Profit/(Loss)',
        'Amount': net_profit
    })
    
    return pd.DataFrame(income_statement)

def generate_balance_sheet(cashbook_data):
    """Generate Balance Sheet from trial balance data."""
    print("Generating Balance Sheet...")
    
    trial_balance = cashbook_data['trial_balance']
    
    # Create balance sheet sections
    balance_sheet = []
    
    # ASSETS
    balance_sheet.append({
        'Account': 'ASSETS',
        'Amount': None
    })
    
    # Add asset accounts
    total_assets = 0
    asset_accounts = trial_balance[trial_balance['Account Type'].str.contains('Asset|Equity', na=False)]
    for _, row in asset_accounts.iterrows():
        net_amount = row['Debit'] - row['Credit']
        total_assets += net_amount
        balance_sheet.append({
            'Account': row['Account'],
            'Amount': net_amount
        })
    
    balance_sheet.append({
        'Account': 'Total Assets',
        'Amount': total_assets
    })
    
    # LIABILITIES
    balance_sheet.append({
        'Account': 'LIABILITIES',
        'Amount': None
    })
    
    # Add liability accounts
    total_liabilities = 0
    liability_accounts = trial_balance[trial_balance['Account Type'].str.contains('Liability', na=False)]
    for _, row in liability_accounts.iterrows():
        net_amount = row['Credit'] - row['Debit']
        total_liabilities += net_amount
        balance_sheet.append({
            'Account': row['Account'],
            'Amount': net_amount
        })
    
    balance_sheet.append({
        'Account': 'Total Liabilities',
        'Amount': total_liabilities
    })
    
    return pd.DataFrame(balance_sheet)

def apply_excel_formatting(ws):
    """Apply consistent formatting with ZAR currency."""
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    # South African Rand format: R # ##0.00
    zar_format = 'R #,##0.00;[Red]-R #,##0.00'
    
    # Format headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Format data cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            if isinstance(cell.value, (int, float)):
                cell.number_format = zar_format
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Add extra width for currency symbol and formatting
        ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 4

def generate_cash_flow_statement(cashbook_data):
    """Generate Cash Flow Statement from transaction data."""
    print("Generating Cash Flow Statement...")
    
    details = cashbook_data['details']
    
    # Group transactions by activity type
    operating_activities = details[details['Account'].str.contains('Income|Expense', na=False)].copy()
    investing_activities = details[details['Account'].str.contains('Asset|Equipment', na=False)].copy()
    financing_activities = details[details['Account'].str.contains('Loan|Capital', na=False)].copy()
    
    # Calculate net cash flows
    operating_cash_flow = operating_activities['amount'].sum()
    investing_cash_flow = investing_activities['amount'].sum()
    financing_cash_flow = financing_activities['amount'].sum()
    
    # Create cash flow statement DataFrame
    cash_flow = []
    
    # Operating Activities
    cash_flow.append({'Category': 'OPERATING ACTIVITIES', 'Amount': None})
    cash_flow.append({'Category': 'Net Cash from Operating Activities', 'Amount': operating_cash_flow})
    
    # Investing Activities
    cash_flow.append({'Category': 'INVESTING ACTIVITIES', 'Amount': None})
    cash_flow.append({'Category': 'Net Cash from Investing Activities', 'Amount': investing_cash_flow})
    
    # Financing Activities
    cash_flow.append({'Category': 'FINANCING ACTIVITIES', 'Amount': None})
    cash_flow.append({'Category': 'Net Cash from Financing Activities', 'Amount': financing_cash_flow})
    
    # Net Change in Cash
    net_change = operating_cash_flow + investing_cash_flow + financing_cash_flow
    cash_flow.append({'Category': 'NET CHANGE IN CASH', 'Amount': net_change})
    
    return pd.DataFrame(cash_flow)

def generate_financial_statements(cashbook_path, output_path=None):
    """Generate and save all financial statements."""
    # Read cashbook data
    cashbook_data = read_cashbook_data(cashbook_path)
    
    # Generate statements
    income_stmt = generate_income_statement(cashbook_data)
    balance_sheet = generate_balance_sheet(cashbook_data)
    cash_flow = generate_cash_flow_statement(cashbook_data)
    
    # Save to Excel if output path provided
    if output_path:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Write each statement to a separate sheet
            income_stmt.to_excel(writer, sheet_name='Income Statement', index=False)
            balance_sheet.to_excel(writer, sheet_name='Balance Sheet', index=False)
            cash_flow.to_excel(writer, sheet_name='Cash Flow', index=False)
            
            # Format each worksheet
            for sheet_name in writer.sheets:
                apply_excel_formatting(writer.sheets[sheet_name])
            
            print(f"\nFinancial statements generated successfully and saved to: {output_path}")
            print("Statements included:")
            print("1. Income Statement")
            print("2. Balance Sheet")
            print("3. Cash Flow Statement")
    
    return {
        'income_statement': income_stmt,
        'balance_sheet': balance_sheet,
        'cash_flow': cash_flow
    }

if __name__ == "__main__":
    # Test the financial statement generation
    import os
    
    base_dir = "/Users/sthwalonyoni/pdf-bank-statement-parser"
    test_input = os.path.join(base_dir, "Annual_Cashbook_FY2024-2025.xlsx")
    test_output = os.path.join(base_dir, "Financial_Statements_Test.xlsx")
    
    generate_financial_statements(test_input, test_output)